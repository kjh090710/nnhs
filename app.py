from flask import Flask, render_template, request, jsonify
import sqlite3
import json
import os
from datetime import datetime
from openpyxl import load_workbook

app = Flask(__name__)
DB_NAME = "students.db"
STUDENTS_FILE = "students.json"


def get_db():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn


def make_student_number(grade, class_no, student_no):
    return f"{int(grade)}{int(class_no):02d}{int(student_no):02d}"


def load_students_from_json():
    if not os.path.exists(STUDENTS_FILE):
        sample_students = [
        ]
        with open(STUDENTS_FILE, "w", encoding="utf-8") as f:
            json.dump(sample_students, f, ensure_ascii=False, indent=2)

    with open(STUDENTS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_students_to_json(students):
    with open(STUDENTS_FILE, "w", encoding="utf-8") as f:
        json.dump(students, f, ensure_ascii=False, indent=2)


def sync_students_to_db(students):
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS students (
            student_number TEXT PRIMARY KEY,
            name TEXT NOT NULL
        )
    """)

    for student in students:
        cur.execute("""
            INSERT OR REPLACE INTO students (student_number, name)
            VALUES (?, ?)
        """, (student["student_number"], student["name"]))

    conn.commit()
    conn.close()


def init_db():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS students (
            student_number TEXT PRIMARY KEY,
            name TEXT NOT NULL
        )
    """)

    # 기존 벌점 개념 제거 → 선도 기록 테이블
    cur.execute("""
        CREATE TABLE IF NOT EXISTS guidances (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_number TEXT NOT NULL,
            student_name TEXT NOT NULL,
            content TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)

    conn.commit()
    conn.close()

    students = load_students_from_json()
    sync_students_to_db(students)


def normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def find_column_index(headers, candidates):
    normalized_headers = [normalize_header(h) for h in headers]
    for i, header in enumerate(normalized_headers):
        if header in candidates:
            return i
    return -1


@app.route("/")
def home():
    return render_template("index.html")


@app.route("/api/student/<student_number>", methods=["GET"])
def get_student(student_number):
    conn = get_db()
    student = conn.execute(
        "SELECT * FROM students WHERE student_number = ?",
        (student_number,)
    ).fetchone()
    conn.close()

    if student:
        return jsonify({
            "success": True,
            "student_number": student["student_number"],
            "name": student["name"]
        })

    return jsonify({
        "success": False,
        "message": "해당 학번의 학생을 찾을 수 없습니다."
    }), 404


@app.route("/api/guidance", methods=["POST"])
def add_guidance():
    data = request.get_json()

    student_number = str(data.get("student_number", "")).strip()
    content = str(data.get("content", "")).strip()

    if not student_number:
        return jsonify({"success": False, "message": "학번을 입력하세요."}), 400

    if not content:
        return jsonify({"success": False, "message": "선도 내용을 입력하세요."}), 400

    conn = get_db()
    student = conn.execute(
        "SELECT * FROM students WHERE student_number = ?",
        (student_number,)
    ).fetchone()

    if not student:
        conn.close()
        return jsonify({"success": False, "message": "학생이 존재하지 않습니다."}), 404

    student_name = student["name"]
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn.execute("""
        INSERT INTO guidances (student_number, student_name, content, created_at)
        VALUES (?, ?, ?, ?)
    """, (student_number, student_name, content, created_at))

    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "message": "선도 내용이 저장되었습니다.",
        "data": {
            "student_number": student_number,
            "student_name": student_name,
            "content": content,
            "created_at": created_at
        }
    })

@app.route("/api/guidance-summary", methods=["GET"])
def get_guidance_summary():
    conn = get_db()

    rows = conn.execute("""
        SELECT
            s.student_number,
            s.name AS student_name,
            COUNT(g.id) AS guidance_count
        FROM students s
        INNER JOIN guidances g
            ON s.student_number = g.student_number
        GROUP BY s.student_number, s.name
        HAVING COUNT(g.id) >= 1
        ORDER BY guidance_count ASC, s.student_number ASC
    """).fetchall()

    conn.close()

    result = []
    for row in rows:
        result.append({
            "student_number": row["student_number"],
            "student_name": row["student_name"],
            "guidance_count": row["guidance_count"]
        })

    return jsonify({
        "success": True,
        "students": result
    })

@app.route("/api/student-guidances/<student_number>", methods=["GET"])
def get_student_guidances(student_number):
    conn = get_db()

    student = conn.execute(
        "SELECT * FROM students WHERE student_number = ?",
        (student_number,)
    ).fetchone()

    if not student:
        conn.close()
        return jsonify({
            "success": False,
            "message": "해당 학번의 학생을 찾을 수 없습니다."
        }), 404

    rows = conn.execute("""
        SELECT id, student_number, student_name, content, created_at
        FROM guidances
        WHERE student_number = ?
        ORDER BY id DESC
    """, (student_number,)).fetchall()

    conn.close()

    guidances = []
    for row in rows:
        guidances.append({
            "id": row["id"],
            "student_number": row["student_number"],
            "student_name": row["student_name"],
            "content": row["content"],
            "created_at": row["created_at"]
        })

    return jsonify({
        "success": True,
        "student_number": student["student_number"],
        "student_name": student["name"],
        "guidance_count": len(guidances),
        "guidances": guidances
    })


@app.route("/api/upload-students-excel", methods=["POST"])
def upload_students_excel():
    if "file" not in request.files:
        return jsonify({"success": False, "message": "파일이 없습니다."}), 400

    file = request.files["file"]

    if file.filename == "":
        return jsonify({"success": False, "message": "선택된 파일이 없습니다."}), 400

    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"success": False, "message": ".xlsx 파일만 업로드 가능합니다."}), 400

    try:
        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return jsonify({"success": False, "message": "엑셀 파일이 비어 있습니다."}), 400

        headers = rows[0]

        grade_idx = find_column_index(headers, {"학년", "grade"})
        class_idx = find_column_index(headers, {"반", "class", "class_no", "class number"})
        student_idx = find_column_index(headers, {"번호", "num", "number", "student_no", "student number"})
        name_idx = find_column_index(headers, {"이름", "name", "학생명"})

        if grade_idx == -1 or class_idx == -1 or student_idx == -1 or name_idx == -1:
            return jsonify({
                "success": False,
                "message": "엑셀 첫 행에 '학년', '반', '번호', '이름' 열이 있어야 합니다."
            }), 400

        student_map = {}

        for row in rows[1:]:
            if not row:
                continue

            try:
                grade = row[grade_idx] if len(row) > grade_idx else None
                class_no = row[class_idx] if len(row) > class_idx else None
                student_no = row[student_idx] if len(row) > student_idx else None
                name = row[name_idx] if len(row) > name_idx else None

                if grade is None or class_no is None or student_no is None or name is None:
                    continue

                name = str(name).strip()
                if not name:
                    continue

                student_number = make_student_number(grade, class_no, student_no)

                # 같은 학생은 1번만 유지
                student_map[student_number] = {
                    "student_number": student_number,
                    "name": name
                }

            except Exception:
                continue

        students = list(student_map.values())
        students.sort(key=lambda x: x["student_number"])

        if not students:
            return jsonify({"success": False, "message": "유효한 학생 데이터가 없습니다."}), 400

        save_students_to_json(students)
        sync_students_to_db(students)

        return jsonify({
            "success": True,
            "message": f"{len(students)}명의 학생 정보가 저장되었습니다."
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"엑셀 처리 중 오류가 발생했습니다: {str(e)}"
        }), 500

@app.route("/api/guidance/<int:guidance_id>", methods=["DELETE"])
def delete_guidance(guidance_id):
    conn = get_db()

    guidance = conn.execute("""
        SELECT id, student_number, student_name, content, created_at
        FROM guidances
        WHERE id = ?
    """, (guidance_id,)).fetchone()

    if not guidance:
        conn.close()
        return jsonify({
            "success": False,
            "message": "삭제할 선도 기록을 찾을 수 없습니다."
        }), 404

    conn.execute("DELETE FROM guidances WHERE id = ?", (guidance_id,))
    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "message": "선도 기록이 삭제되었습니다.",
        "data": {
            "id": guidance["id"],
            "student_number": guidance["student_number"],
            "student_name": guidance["student_name"]
        }
    })

if __name__ == "__main__":
    init_db()
    app.run(debug=True)